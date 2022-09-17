import sys
import pandas as pd
import lxml
import numpy as np
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill, colors
from pathlib import Path

# git bash 상에서 실행하기 위해 sys.argv 활용
# 문제 : git bash에서 year를 설정할 시 데이터 추출의 year는 statiz_background.py, 엑셀 파일 이름 설정의 year는 main.py에 적용되는듯?
# record_site 및 kt_site를 구동할 수 있는 하나의 메서드로 새로 만들어야? 복사 파일 활용하여 진행해볼 필요
year = sys.argv[1]

# 타격기본, 타격확장, 투수기본, 투수확장, 투수구속의 첫 페이지를 딕셔너리 형태로 정리
record_site = {'타격기본':f'http://www.statiz.co.kr/stat.php?opt=0&sopt=0&re=0&ys={year}&ye={year}&se=0&te=&tm=&ty=0&qu=auto&po=0&as=&ae=&hi=&un=&pl=&da=1&o1=WAR_ALL_ADJ&o2=TPA&de=1&lr=0&tr=&cv=&ml=1&sn=30&si=&cn=',
        '타격확장':f'http://www.statiz.co.kr/stat.php?opt=0&sopt=0&re=0&ys={year}&ye={year}&se=0&te=&tm=&ty=0&qu=all&po=0&as=&ae=&hi=&un=&pl=&da=2&o1=WRCPLUS&o2=WAR_ALL&de=1&lr=0&tr=&cv=&ml=1&sn=30&si=&cn=',
        '투수기본':f'http://www.statiz.co.kr/stat.php?opt=0&sopt=0&re=1&ys={year}&ye={year}&se=0&te=&tm=&ty=0&qu=auto&po=0&as=&ae=&hi=&un=&pl=&da=1&o1=WAR&o2=OutCount&de=1&lr=0&tr=&cv=&ml=1&sn=30&si=&cn=',
        '투수확장':f'http://www.statiz.co.kr/stat.php?opt=0&sopt=0&re=1&ys={year}&ye={year}&se=0&te=&tm=&ty=0&qu=all&po=0&as=&ae=&hi=&un=&pl=&da=2&o1=FIP&o2=WAR&de=0&lr=0&tr=&cv=&ml=1&sn=30&si=&cn=',
        '투수구속':f'http://www.statiz.co.kr/stat.php?opt=0&sopt=0&re=1&ys={year}&ye={year}&se=0&te=&tm=&ty=0&qu=all&po=0&as=&ae=&hi=&un=&pl=&da=14&o1=FVval&de=1&o2=WAR&lr=0&tr=&cv=&ml=1&sn=30&si=&cn='
        }
# 스탯티즈에서는 KIA및 KT가 똑같이 'K'로 나오므로, KT선수를 구분해내기 위해 해당 페이지 딕셔너리로 정리
kt_site = {'타격':f'http://www.statiz.co.kr/stat.php?mid=stat&re=0&ys={year}&ye={year}&se=0&te=kt&tm=&ty=0&qu=auto&po=0&as=&ae=&hi=&un=&pl=&da=1&o1=WAR_ALL_ADJ&o2=TPA&de=1&lr=0&tr=&cv=&ml=1&sn=30&pa=0&si=&cn=',
        '투수':f'http://www.statiz.co.kr/stat.php?opt=0&sopt=0&re=1&ys={year}&ye={year}&se=0&te=kt&tm=&ty=0&qu=auto&po=0&as=&ae=&hi=&un=&pl=&da=1&o1=WAR&o2=OutCount&de=1&lr=0&tr=&cv=&ml=1&sn=30&si=&cn='}

# 클래스 형태로 정리
class Statiz:
    def __init__(self):
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    
    # 일단 미활용 → 다른 속성/메서드 내 지역함수 형태로 활용
    def table_change(self, df):
        df = df.droplevel(0, axis=1)
        # pop() 메서드나 del[] 속성으로 제거해도 정렬에 있는 것이 중복으로 발생하므로, 굳이 제거하지 않고 진행
        return df
    
    # 일단 미활용 → 다른 속성/메서드 내 지역함수 형태로 활용
    def remove_records(self, df):
        df = df.query("이름 != '이름'")
        return df
    
    # 일단 미활용 → 다른 속성/메서드 내 지역함수 형태로 활용
    def numeric(self, df):
        for col in df.columns.tolist():
            df[col] = pd.to_numeric(df[col], errors='ignore')
        return df

    # @property를 통해 속성으로 바꿔주긴 하는데.. 하나의 메서드로 합칠 수는 없나?
    # 일단 다 만들어놓고 리팩토링을 생각해보자
    # 두 가지의 속성을 하나로 합쳐서 하나의 메서드로 합치는 것은 불가능할듯? - 타격기본, 타격확장, 투수기본, 투수확장, 투수구속으로 구분
    @property
    def hit_basic(self):
        self.driver.get(record_site['타격기본'])
        # 각 페이지들의 DataFrame 변환본을 리스트에 모두 할당하기 위해 빈 리스트 생성
        hit_basic_df_list = []
        while True:
            page = self.driver.page_source
            # BeautifulSoup 크롤링 형태 대신 Pandas의 read_html 함수 활용 → 두 개의 DataFrame을 반환하므로 두 번째 DataFrame만 선택
            # BeautifulSoup때도 발생했는데, 아마 스탯티즈가 두 개의 DataFrame을 합쳐놓은 형태로 사이트가 구성된듯
            df = pd.read_html(page)[1]
            hit_basic_df_list.append(df)
            try:
                # Selenium의 find_element() 메서드 활용하여 사이트 직접 컨트롤
                # By.TAG_NAME 및 By.LINK_TEXT 전달인자 활용
                html = self.driver.find_element(By.TAG_NAME, 'html')
                next_link = self.driver.find_element(By.LINK_TEXT, '다음')
                html.send_keys(Keys.END)
                time.sleep(0.3)
                next_link.click()
                time.sleep(0.3)
            except:
                print('수집 종료')
                break
        hit_basic_df = pd.concat(hit_basic_df_list)
        # 원래는 여기까지만 하나의 속성으로 만들고 밑 내용은 분리시켰는데, 문제가 발생하여 하나의 속성으로 합치기
        # 필요한 columns 정리
        hit_basic_cols = ['순', '이름', '팀', 'G', '타석', '타수', '득점', '안타', '2타', '3타', '홈런',
        '루타', '타점', '도루', '도실', '볼넷', '사구', '고4', '삼진', '병살', '희타', '희비',
        '타율', '출루', '장타', 'OPS', 'wOBA', 'wRC+', 'WAR*', 'WPA']
        hit_basic_df = (hit_basic_df
        # 활용하려던 클래스 메서드 대신 람다식 형태로 활용
        .pipe(lambda df: df.droplevel(0, axis=1))
        # 10명 씩 세 묶음으로 되어 있어 불필요한 내용이 포함되므로, 해당 내용 람다식으로 제거
        .pipe(lambda df: df.query("이름 != '이름'"))
        .loc[:, hit_basic_cols]
        )
        # 정렬 기준으로 인해 같은 열이 두 개 생기게 되는 경우가 있어 Series로 추출한 후 DataFrame에 assign 진행
        hit_basic_sep = hit_basic_df.loc[:, 'WAR*'].iloc[:, 0]
        hit_basic_df = (hit_basic_df
        .drop(['WAR*', '순'], axis=1)
        .assign(**{'WAR*':hit_basic_sep})
        )
        return hit_basic_df
    
    # 타격기본과 같이 타격확장, 투수기본, 투수확장, 투수구속도 @property 활용하여 같은 형태로 속성 생성
    @property
    def hit_expand(self):
        self.driver.get(record_site['타격확장'])
        hit_expand_df_list = []
        while True:
            page = self.driver.page_source
            df = pd.read_html(page)[1]
            hit_expand_df_list.append(df)
            try:
                html = self.driver.find_element(By.TAG_NAME, 'html')
                next_link = self.driver.find_element(By.LINK_TEXT, '다음')
                html.send_keys(Keys.END)
                time.sleep(0.3)
                next_link.click()
                time.sleep(0.3)
            except:
                print('수집 종료')
                break
        hit_expand_df = pd.concat(hit_expand_df_list)
        hit_expand_cols = ['순', '이름', '팀', '타석', 'HR%', 'BB%', 'K%', 'BB/K', 'IsoP', 'IsoD',
        'BABIP', 'Spd', 'PSN', 'wRC+']
        hit_expand_df = (hit_expand_df
        .pipe(lambda df: df.droplevel(0, axis=1))
        .pipe(lambda df: df.query("이름 != '이름'"))
        .loc[:, hit_expand_cols]
        )
        hit_expand_sep = hit_expand_df.loc[:, 'wRC+'].iloc[:, 0]
        hit_expand_df = (hit_expand_df
        .drop(['wRC+', '순'], axis=1)
        .assign(**{'wRC+':hit_expand_sep})
        )
        return hit_expand_df
    
    @property
    def pitch_basic(self):
        self.driver.get(record_site['투수기본'])
        pitch_basic_df_list = []
        while True:
            page = self.driver.page_source
            df = pd.read_html(page)[1]
            pitch_basic_df_list.append(df)
            try:
                html = self.driver.find_element(By.TAG_NAME, 'html')
                next_link = self.driver.find_element(By.LINK_TEXT, '다음')
                html.send_keys(Keys.END)
                time.sleep(0.3)
                next_link.click()
                time.sleep(0.3)
            except:
                print('수집 종료')
                break
        pitch_basic_df = pd.concat(pitch_basic_df_list)
        pitch_basic_cols = ['순', '이름', '팀', '출장', '완투', '완봉', '선발', '승', '패', '세', '홀드', '이닝', '실점',
        '자책', '타자', '안타', '2타', '3타', '홈런', '볼넷', '고4', '사구', '삼진', '보크', '폭투',
        'ERA', 'FIP', 'WHIP', 'ERA+', 'FIP+', 'WAR', 'WPA']
        pitch_basic_df = (pitch_basic_df
        .pipe(lambda df: df.droplevel(0, axis=1))
        .pipe(lambda df: df.query("이름 != '이름'"))
        .loc[:, pitch_basic_cols]
        )
        pitch_basic_sep = pitch_basic_df.loc[:, 'WAR'].iloc[:, 0]
        pitch_basic_df = (pitch_basic_df
        .drop(['WAR', '순'], axis=1)
        .assign(**{'WAR':pitch_basic_sep})
        )
        return pitch_basic_df
    
    @property
    def pitch_expand(self):
        self.driver.get(record_site['투수확장'])
        pitch_expand_df_list = []
        while True:
            page = self.driver.page_source
            df = pd.read_html(page)[1]
            pitch_expand_df_list.append(df)
            try:
                html = self.driver.find_element(By.TAG_NAME, 'html')
                next_link = self.driver.find_element(By.LINK_TEXT, '다음')
                html.send_keys(Keys.END)
                time.sleep(0.3)
                next_link.click()
                time.sleep(0.3)
            except:
                print('수집 종료')
                break
        pitch_expand_df = pd.concat(pitch_expand_df_list)
        pitch_expand_cols = ['순', '이름', '팀', '출장', '이닝', 'ERA', 'FIP', 'K/9', 'BB/9', 'K/BB',
        'HR/9', 'K%', 'BB%', 'K-BB%', 'PFR', 'BABIP', 'LOB%', '타율', '출루율',
        '장타율', 'OPS', 'WHIP', 'WHIP+', '투구', 'IP/G', 'P/G', 'P/IP', 'P/PA',
        'CYP']
        pitch_expand_df = (pitch_expand_df
        .pipe(lambda df: df.droplevel(0, axis=1))
        .pipe(lambda df: df.query("이름 != '이름'"))
        .loc[:, pitch_expand_cols]
        )
        pitch_expand_sep = pitch_expand_df.loc[:, 'FIP'].iloc[:, 0]
        pitch_expand_df = (pitch_expand_df
        .drop(['FIP', '순'], axis=1)
        .assign(**{'FIP':pitch_expand_sep})
        )
        return pitch_expand_df
    
    @property
    def pitch_speed(self):
        self.driver.get(record_site['투수구속'])
        pitch_speed_df_list = []
        while True:
            page = self.driver.page_source
            df = pd.read_html(page)[1]
            pitch_speed_df_list.append(df)
            try:
                html = self.driver.find_element(By.TAG_NAME, 'html')
                next_link = self.driver.find_element(By.LINK_TEXT, '다음')
                html.send_keys(Keys.END)
                time.sleep(0.3)
                next_link.click()
                time.sleep(0.3)
            except:
                print('수집 종료')
                break
        pitch_speed_df = pd.concat(pitch_speed_df_list)
        # 투수구속의 경우 최대구속을 추출하여야 하므로, 일단 먼저 numeric 지역함수 생성하여 숫자형태로 변환
        def numeric(df):
            for col in df.columns.tolist():
                df[col] = pd.to_numeric(df[col], errors='ignore')
            return df
        # 스탯티즈 상에는 원래 columns가 2-level로 구성되어 있어 이렇게 정리하면 두 종류가 생성
        # ex. 직구 → 직구구속 및 직구구사율이 '직구'라는 하나의 열 이름으로 동시에 발생
        # 아래쪽 pitch_speed_df.columns를 직접 변경하여 분리
        pitch_speed_cols = ['순', '이름', '팀', '출장', '이닝', '직구', '슬라', '커브', '첸졉', '스플', '싱커',
        '너클', '기타']
        pitch_speed_df = (pitch_speed_df
        .pipe(lambda df: df.droplevel(0, axis=1))
        .pipe(lambda df: df.query("이름 != '이름'"))
        .loc[:, pitch_speed_cols]
        )
        pitch_speed_df.columns = ['순', '이름', '팀', '출장', '이닝', '직구구속', '직구구속2', '직구구사', '슬라구속', '슬라구사', '커브구속', '커브구사',
        '첸졉구속', '첸졉구사', '스플구속', '스플구사', '싱커구속', '싱커구사', '너클구속', '너클구사', '기타구속', '기타구사']
        pitch_speed_df = (pitch_speed_df
        .drop('직구구속2', axis=1)
        .pipe(numeric)
        .assign(최고구속=lambda df: df.loc[:, ['직구구속', '슬라구속', '커브구속', '첸졉구속', '스플구속', '싱커구속', '너클구속', '기타구속']].max(axis=1))
        .loc[:, ['순', '이름', '팀', '출장', '이닝', '최고구속', '직구구속', '슬라구속', '커브구속', '첸졉구속', '스플구속', '싱커구속', '너클구속', '기타구속', 
        '직구구사', '슬라구사', '커브구사', '첸졉구사', '스플구사', '싱커구사', '너클구사', '기타구사']]
        )
        return pitch_speed_df
    
    # 팀에서 KIA와 KT를 구분하기 위해 kt_team_member_hit/pitch 속성을 별도로 생성
    # 속성의 결과값은 선수들 이름이 포함된 리스트
    @property
    def kt_team_member_hit(self):
        self.driver.get(kt_site['타격'])
        kt_df_list_hit = []
        while True:
            page = self.driver.page_source
            df = pd.read_html(page)[1]
            kt_df_list_hit.append(df)
            try:
                html = self.driver.find_element(By.TAG_NAME, 'html')
                next_link = self.driver.find_element(By.LINK_TEXT, '다음')
                html.send_keys(Keys.END)
                time.sleep(0.3)
                next_link.click()
                time.sleep(0.3)
            except:
                print('수집 종료')
                break
        kt_team_member_hit = (pd.concat(kt_df_list_hit)
        .pipe(lambda df: df.droplevel(0, axis=1))
        ['이름']
        .tolist()
        )
        return kt_team_member_hit
    
    @property
    def kt_team_member_pitch(self):
        self.driver.get(kt_site['투수'])
        kt_df_list_pitch = []
        while True:
            page = self.driver.page_source
            df = pd.read_html(page)[1]
            kt_df_list_pitch.append(df)
            try:
                html = self.driver.find_element(By.TAG_NAME, 'html')
                next_link = self.driver.find_element(By.LINK_TEXT, '다음')
                html.send_keys(Keys.END)
                time.sleep(0.3)
                next_link.click()
                time.sleep(0.3)
            except:
                print('수집 종료')
                break
        kt_team_member_pitch = (pd.concat(kt_df_list_pitch)
        .pipe(lambda df: df.droplevel(0, axis=1))
        ['이름']
        .tolist()
        )
        return kt_team_member_pitch
    
    # 추출한 속성들을 성격에 맞게 병합해줘야(ex. 타격기본 + 타격확장 = 타자)
    def merge_hit(self):
        # numeric 지역함수 새로 생성하여 값들을 숫자로 변환
        def numeric(df):
            for col in df.columns.tolist():
                df[col] = pd.to_numeric(df[col], errors='ignore')
            return df
        # merge() 메서드를 이용해 병합 진행 → on 파라미터를 최대한 포함하여 동명이인이 합쳐지지 않도록 주의
        self.hit_df = (self.hit_basic
        .merge(self.hit_expand, on=['이름', '팀', '타석', 'wRC+'], how='inner')
        .pipe(numeric)
        .fillna(0)
        )
        hit_columns = ['이름', '연도', '소속', '포지션', 'G', '타석', '타수', '득점', '안타', '2타', '3타', '홈런', '루타', '타점',
        '도루', '도실', '볼넷', '사구', '고4', '삼진', '병살', '희타', '희비', '타율', '출루', '장타',
        'OPS', 'wOBA', 'wRC+', 'WPA', 'WAR*', 'HR%', 'BB%', 'K%', 'BB/K',
        'IsoP', 'IsoD', 'BABIP', 'Spd', 'PSN']
        self.hit_df = (self.hit_df
        # 팀이 '18삼CF' 형태로 연도, 팀, 포지션이 합쳐져있으므로 해당 내용 분리 및 각 열에 할당 진행
        # 연도의 경우 str 속성의 slice 활용하여 앞 두 글자 추출
        .assign(연도=lambda df: '20' + df['팀'].str.slice(0, 2),
        # 소속의 경우 str 속성의 extract 활용하여 정규표현식 형태로 진행
        소속=lambda df: df['팀'].str.extract('(L|키|K|S|N|한|삼|롯|두|넥)')
        # 그리고 각 내용들을 실제 팀 이름으로 replace
        .replace({'L':'LG', '키':'넥센/키움', 'S':'SK/SSG', 'N':'NC', '한':'한화', '삼':'삼성', '롯':'롯데', '두':'두산', '넥':'넥센/키움'}),
        # 포지션 역시 str 속성의 extract 활용
        포지션=lambda df: df['팀'].str.extract('(LF|CF|RF|1B|2B|3B|SS|C|DH|P)'))
        # 위에서 추출한 kt_team_member_hit/pitch 리스트 활용, np.where() 함수 이용
        .assign(소속=lambda df: np.where(df['이름'].isin(self.kt_team_member_hit), 'KT', df['소속']))
        .loc[:, hit_columns]
        )
        return self.hit_df
    
    # 마찬가지 형식으로 추출한 속성들을 성격에 맞게 병합 진행(ex. 투수기본 + 투수확장 + 투수구속 = 투수)
    def merge_pitch(self):
        def numeric(df):
            for col in df.columns.tolist():
                df[col] = pd.to_numeric(df[col], errors='ignore')
            return df
        self.pitch_df = (self.pitch_basic
        .merge(self.pitch_expand, on=['이름', '팀', '출장', '이닝', 'ERA', 'FIP', 'WHIP'], how='inner')
        .pipe(numeric)
        .merge(self.pitch_speed, on=['이름', '팀', '출장', '이닝'], how='inner')
        .fillna(0)
        )
        pitch_columns = ['이름', '연도', '소속', '출장', '완투', '완봉', '선발', '승', '패', '세', '홀드', '이닝', '실점',
        '자책', '타자', '안타', '2타', '3타', '홈런', '볼넷', '고4', '사구', '삼진', '보크', '폭투',
        'ERA', 'FIP', 'WHIP', 'ERA+', 'FIP+', 'WPA', 'WAR', 'K/9', 'BB/9',
        'K/BB', 'HR/9', 'K%', 'BB%', 'K-BB%', 'PFR', 'BABIP', 'LOB%', '타율',
        '출루율', '장타율', 'OPS', 'WHIP+', '투구', 'IP/G', 'P/G', 'P/IP', 'P/PA',
        'CYP', '순', '최고구속', '직구구속', '슬라구속', '커브구속', '첸졉구속', '스플구속', '싱커구속',
        '너클구속', '기타구속', '직구구사', '슬라구사', '커브구사', '첸졉구사', '스플구사', '싱커구사', '너클구사',
        '기타구사']
        self.pitch_df = (self.pitch_df
        .assign(연도=lambda df: '20' + df['팀'].str.slice(0, 2),
        소속=lambda df: df['팀'].str.extract('(L|키|K|S|N|한|삼|롯|두|넥)')
        .replace({'L':'LG', '키':'넥센/키움', 'S':'SK/SSG', 'N':'NC', '한':'한화', '삼':'삼성', '롯':'롯데', '두':'두산', '넥':'넥센/키움'}),
        포지션=lambda df: df['팀'].str.extract('(LF|CF|RF|1B|2B|3B|SS|C|DH|P)'))
        .assign(소속=lambda df: np.where(df['이름'].isin(self.kt_team_member_pitch), 'KT', df['소속']))
        .loc[:, pitch_columns]
        )
        return self.pitch_df
    
    # 엑셀로 export 하는 메서드 생성 진행
    def export_excel(self, year):
        file_path = Path.home() / 'statiz_project_3.1' / 'result_file' / f'Statiz_{year}.xlsx'
        excel_file = pd.ExcelWriter(file_path, engine='openpyxl')
        # merge_hit() 메서드와 merge_pitch() 메서드를 각각 Pandas의 to_excel() 메서드로 엑셀 추출
        self.merge_hit().to_excel(excel_file, sheet_name='타자', index=False)
        self.merge_pitch().to_excel(excel_file, sheet_name='투수', index=False)
        excel_file.save()
        # openpyxl 활용하여 셀 형태 조정
        excel_file = openpyxl.load_workbook(file_path)
        for sheet_name in ['타자', '투수']:
            sheet = excel_file[sheet_name]
            # Font() 함수 활용하여 bold 지정
            bold_font = Font(bold=True)
            # Alignment() 함수 활용하여 중앙정렬 지정
            center_alignment = Alignment(horizontal='center', vertical='center')
            # PatternFill() 함수 활용하여 셀 배경색을 노란색으로 지정
            yellow_fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
            for i in range(1, sheet.max_column + 1):
                sheet.cell(1, i).font = bold_font
                sheet.cell(1, i).alignment = center_alignment
                sheet.cell(1, i).fill = yellow_fill
        excel_file.save(file_path)


